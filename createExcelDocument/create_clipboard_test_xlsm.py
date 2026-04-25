"""
Build email_contents/clipboard_test.xlsm — minimal workbook to debug Open File Location + VBA.

Matches the main orders workbook: Open File Location cells use in-workbook # hyperlinks only;
the real file:/// URI lives in hidden column AC (29). VBA uses
Workbook_SheetFollowHyperlink and reads the URI from column 29.

Requires Windows + Excel + pywin32 (same as macro template).

Usage (from python_files):
  python createExcelDocument/create_clipboard_test_xlsm.py

Writes:
  <BASE_DIR>/email_contents/clipboard_test.xlsm
Uses the same clipboard ini as the main exporter (default: python_files/excel_clipboard_launch.ini).
"""
from __future__ import annotations

import importlib.util
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

_PYTHON_FILES = Path(__file__).resolve().parent.parent
if str(_PYTHON_FILES) not in sys.path:
    sys.path.insert(0, str(_PYTHON_FILES))

from shared.project_paths import ensure_base_dir_in_environ
from shared.settings_store import apply_runtime_settings_from_json

apply_runtime_settings_from_json()

# Same as createExcelDocument.COPY_PATH_URI_COL and VBA COL_FILE_URI
COL_FILE_URI = 29


def _load_macro_template():
    mpath = Path(__file__).resolve().parent / "macro_template.py"
    spec = importlib.util.spec_from_file_location("_macro_template", mpath)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {mpath}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def main() -> int:
    email_contents = ensure_base_dir_in_environ() / "email_contents"
    email_contents.mkdir(parents=True, exist_ok=True)

    ini_path = Path(
        os.getenv("EXCEL_CLIPBOARD_INI_PATH", str(_PYTHON_FILES / "excel_clipboard_launch.ini"))
    ).expanduser().resolve()
    test_xlsm = email_contents / "clipboard_test.xlsm"

    macro = _load_macro_template()
    print(f"Building macro shell: {test_xlsm}")
    if not macro.build_macro_template_file(test_xlsm):
        print(
            "Excel COM build failed. Fix Excel/pywin32, or copy orders_template.xlsm to "
            "clipboard_test.xlsm and re-run this script (openpyxl step only will run if file exists)."
        )
        return 1

    wb = load_workbook(test_xlsm, keep_vba=True)
    ws = wb.active
    ws.title = "Orders"

    # Minimal grid: B = Open File Location header; AC (29) = plain file URI; B2 = internal link only.
    ws["A1"] = "Note"
    ws["B1"] = "Open File Location"
    ws["A2"] = "Click B2: Explorer opens with this ini file selected."
    file_uri = ini_path.resolve().as_uri()
    ws.cell(row=2, column=COL_FILE_URI, value=file_uri)
    ws.column_dimensions[get_column_letter(COL_FILE_URI)].hidden = True

    cell = ws["B2"]
    cell.value = "Open File Location"
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location="Orders!$B$2")
    cell.font = Font(name="Calibri", color="0563C1", underline="single")

    ws["AA1"] = str(ini_path.resolve())
    ws.column_dimensions["AA"].hidden = True

    wb.save(test_xlsm)
    print(f"Wrote {test_xlsm}")
    if not ini_path.is_file():
        print(
            f"WARNING: {ini_path} does not exist yet. Run createExcelDocument once, "
            "or create the ini manually, then test again."
        )
    else:
        print(f"Using ini: {ini_path}")
    print("Open clipboard_test.xlsm, enable macros, click the blue Open File Location in B2.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
