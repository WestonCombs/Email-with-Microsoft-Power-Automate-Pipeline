import importlib.util
import json
import os
from collections.abc import Callable
from copy import copy
import re
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, unquote
from urllib.request import url2pathname

_PYTHON_FILES_DIR = Path(__file__).resolve().parent.parent
if str(_PYTHON_FILES_DIR) not in sys.path:
    sys.path.insert(0, str(_PYTHON_FILES_DIR))

from shared.stdio_utf8 import configure_stdio_utf8, console_safe_text

configure_stdio_utf8()

from shared.settings_store import apply_runtime_settings_from_json
from shared.runLogger import is_debug
from giftcardInvoiceLink.link_store import (
    gift_order_link_label,
    load_edges,
    links_path_for_project_root,
    normalized_order_number,
    stable_record_key,
)
from htmlHandler.tracking_hrefs import MULTIPLE_TRACKING_LINKS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from proofOfDelivery.pod_data import (
    AUTOMATION_HUB_CATEGORY,
    AUTOMATION_HUB_STATUS_LABEL,
    POD_CATEGORY,
    automation_hub_record,
    is_automation_hub_record,
    is_pod_record,
    load_excel_records,
)

apply_runtime_settings_from_json()


def _use_legacy_category_row_colors() -> bool:
    """When true, row background uses email category (Invoice/Shipped/…); default is order-number bands."""
    v = (os.getenv("EXCEL_LEGACY_CATEGORY_ROW_COLORS") or "").strip().lower()
    return v in ("1", "true", "yes", "on")


_base_dir_raw = os.getenv("BASE_DIR")
if not _base_dir_raw:
    raise ValueError(
        'BASE_DIR is not set — expected automatic detection from the "python_files" folder layout.'
    )

PROJECT_ROOT = Path(_base_dir_raw).expanduser().resolve()
_JSON_DIR = PROJECT_ROOT / "email_contents" / "json"
JSON_PATH = str(_JSON_DIR / "results.json")
# Written after each full Orders rebuild; used to skip redundant rebuilds when only Shipping Status refresh is needed.
EXCEL_BUILD_STATE_PATH = _JSON_DIR / "excel_build_state.json"

# Hidden on Orders; VBA reads this before ThisWorkbook.Path (works if Path is empty).
CLIPBOARD_INI_CELL = "AA1"
ACTION_ROW = 1
HEADER_ROW = 2
DATA_START_ROW = 3
FREEZE_PANES_CELL = "B3"
# Hidden: plain-text file:/// URI per row. Open File Location cells use internal # links only
# (Excel hyperlink events cannot cancel file:// navigation).
COPY_PATH_URI_COL = 29  # column AC — keep equal to COL_FILE_URI in macro_template.py
# Hidden full tracking URLs: column 30 + slot-1. Must match VBA COL_TRACK_URI_* in macro_template.py
TRACKING_URI_COL_START = 30  # AD
# One hidden column per visible tracking slot (Track 1 … Track N).
TRACKING_LINK_VISIBLE_SLOTS = 15
TRACKING_URI_SLOT_COUNT = TRACKING_LINK_VISIBLE_SLOTS  # AD through AR (30–44)

# Hidden tracking numbers (plain text IDs) for VBA — must match COL_TRACK_NUM_* in macro_template.py
TRACKING_NUMBER_COL_START = TRACKING_URI_COL_START + TRACKING_URI_SLOT_COUNT  # 45 (AS)
TRACKING_NUMBER_SLOT_COUNT = TRACKING_LINK_VISIBLE_SLOTS
# Hidden 1/0: ID also found on a classified tracking URL (cross-check with link pipeline). BE–BU (60–74).
TRACKING_NUM_CONFIRM_COL_START = TRACKING_NUMBER_COL_START + TRACKING_NUMBER_SLOT_COUNT  # 60
TRACKING_NUM_CONFIRM_SLOT_COUNT = TRACKING_LINK_VISIBLE_SLOTS

# Optional .env overrides (absolute or relative paths expand from user / cwd).
def _optional_path(env_name: str, default: Path) -> Path:
    raw = os.getenv(env_name)
    if raw:
        return Path(raw).expanduser().resolve()
    return default


def _default_orders_template_path() -> Path:
    return PROJECT_ROOT / "email_contents" / "orders_template.xlsm"


ORDERS_TEMPLATE_PATH = _optional_path(
    "EXCEL_TEMPLATE_PATH", _default_orders_template_path()
)
# VBA reads absolute path from cell AA1; ini can live outside email_contents.
CLIPBOARD_LAUNCH_INI_PATH = _optional_path(
    "EXCEL_CLIPBOARD_INI_PATH", _PYTHON_FILES_DIR / "excel_clipboard_launch.ini"
)
EXCEL_PATH = str(
    _optional_path("EXCEL_OUTPUT_PATH", PROJECT_ROOT / "email_contents" / "orders.xlsm")
)

# Launched from Excel VBA (VIEWER= in excel_clipboard_launch.ini) for the "View Tracking Links" column.
TRACKING_VIEWER_SCRIPT = _PYTHON_FILES_DIR / "trackingLinkViewer" / "tracking_link_viewer.py"
# TRACKING_NUMBERS_VIEWER= in ini — "View Tracking Numbers" / "order" aggregate modes.
TRACKING_NUMBERS_VIEWER_SCRIPT = _PYTHON_FILES_DIR / "trackingNumbersViewer" / "tracking_numbers_viewer.py"
# TRACKING_STATUS_VIEWER= — "Shipping Status" column click → 17TRACK status viewer.
TRACKING_STATUS_VIEWER_SCRIPT = _PYTHON_FILES_DIR / "trackingNumbersViewer" / "tracking_status_viewer.py"
# Launched from VBA (GIFTCARD_LINK= in ini) when the user follows an ``Invoice link`` cell.
GIFT_INVOICE_LINK_SCRIPT = _PYTHON_FILES_DIR / "giftcardInvoiceLink" / "gift_invoice_link_workflow.py"
POD_WORKFLOW_SCRIPT = _PYTHON_FILES_DIR / "proofOfDelivery" / "pod_workflow.py"


def _resolve_excel_output_path(using_template: bool) -> str:
    """Use .xlsx unless we loaded a real .xlsm template (VBA). Plain openpyxl .xlsm saves break Excel."""
    p = Path(EXCEL_PATH)
    if using_template:
        if p.suffix.lower() != ".xlsm":
            out = str(p.with_suffix(".xlsm"))
            print(
                f"Note: Macro template in use - saving to '{out}' (.xlsm) instead of '{p.name}'."
            )
            return out
        return str(p)
    if p.suffix.lower() == ".xlsm":
        out = str(p.with_suffix(".xlsx"))
        print(
            f"Note: No orders_template.xlsm — saving to '{out}' (.xlsx). "
            "Saving plain openpyxl output as .xlsm is invalid and Excel will not open it."
        )
        return out
    return str(p)


def _macro_template_module():
    """Load sibling macro_template.py without treating this folder as a package."""
    mpath = Path(__file__).resolve().parent / "macro_template.py"
    spec = importlib.util.spec_from_file_location("_email_sorter_macro_template", mpath)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load macro helper: {mpath}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# Column layout: order identity, then Open File Location + View PDF (adjacent, right after Email),
# then financials and shipping. Open File Location uses in-sheet # links; hidden column AC holds the
# file URI for VBA. View PDF / View HTML (debug) are file:/// hyperlinks.
# Hairline vertical borders after the last file-link column and after Tax Paid (section breaks).
def _section_boundary_keys(column_keys: list[str]) -> tuple[str, str]:
    view_end = (
        "html_source_link" if "html_source_link" in column_keys else "source_file_link"
    )
    return (view_end, "tax_paid")


# Tracking URLs + tracking numbers: hidden columns for VBA; debug-only link columns for viewers.
_COLUMN_ORDER_SUFFIX = [
    "open_tracking_list",
    "open_tracking_numbers_web",
    "open_tracking_numbers_order",
]


def _column_order_keys() -> list[str]:
    keys = [
        "email_category",
        "order_number",
        "purchase_datetime",
        "company",
        "email",
        "copy_file_path",
        "source_file_link",
    ]
    if is_debug():
        keys.append("html_source_link")
    keys.extend(
        [
            "total_amount_paid",
            "tax_paid",
            "gift_invoice_action",
        ]
    )
    keys.append("tracking_quick_status")
    keys.extend(_COLUMN_ORDER_SUFFIX)
    return keys

COLUMN_HEADERS = {
    "email_category":        "Category",
    "order_number":          "Order Number",
    "gift_invoice_action":   "Invoice Link",
    "purchase_datetime":     "Purchase Date",
    "company":               "Company",
    "email":                 "Email",
    "total_amount_paid":     "Total Paid",
    "tax_paid":              "Tax Paid",
    "tracking_quick_status": "Shipping Status",
    "open_tracking_list":    "View Tracking Links",
    "open_tracking_numbers_web": "View Tracking Numbers",
    "open_tracking_numbers_order": "View Tracking Numbers (All For Order)",
    "copy_file_path":        "Open File Location",
    "source_file_link":      "View PDF",
    "html_source_link":      "View HTML",
}

# Recognize older workbooks before we renamed the column / merged with "View shipping status".
SHIPPING_STATUS_HEADER_ALIASES = frozenset(
    {
        COLUMN_HEADERS["tracking_quick_status"],
        "Shipping summary",
        "View shipping status",
        "View Shipping Status",
    }
)

HYPERLINK_FONT = Font(name="Calibri", color="0563C1", underline="single")
HEADER_FILL    = PatternFill("solid", fgColor="2F5597")
HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Calibri")
CELL_FONT      = Font(name="Calibri")
CENTER_ALIGN   = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN     = Alignment(horizontal="left",   vertical="center")

CATEGORY_FILLS = {
    "Invoice":   PatternFill("solid", fgColor="E3F2FD"),  # light sky
    "Shipped":   PatternFill("solid", fgColor="FFF3E0"),  # light peach
    "Delivered": PatternFill("solid", fgColor="E8F5E9"),  # light mint
    "Gift Card": PatternFill("solid", fgColor="FFF9C4"),  # light yellow
    "Unknown":   PatternFill("solid", fgColor="F3E5F5"),  # light purple
    AUTOMATION_HUB_CATEGORY: PatternFill("solid", fgColor="FFB347"),  # vivid amber
}

# Order-number mode: alternate by distinct order id (file order); gift rows stay yellow; no order → purple.
# Green 200 + blue 200 (Material): stronger hue separation than the old green-50 / pale-blue pair.
ORDER_BAND_FILLS = (
    PatternFill("solid", fgColor="A5D6A7"),  # green 200
    PatternFill("solid", fgColor="90CAF9"),  # blue 200
)
UNKNOWN_ORDER_FILL = CATEGORY_FILLS["Unknown"]

HAIR_SIDE = Side(style="hair", color="000000")
AUTOMATION_HUB_FONT = Font(name="Calibri", color="5A2500", bold=True)
AUTOMATION_HUB_LINK_FONT = HYPERLINK_FONT


def load_json(path: str) -> list[dict]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def current_excel_build_debug_flag() -> str:
    """``'1'`` or ``'0'`` — same rule as :func:`runLogger.is_debug` / ``DEBUG_MODE`` in ``.env``."""
    return "1" if is_debug() else "0"


def read_excel_build_debug_mode() -> str | None:
    """
    Return the ``DEBUG_MODE`` flag (``'1'`` / ``'0'``) last recorded for a full workbook layout build,
    or ``None`` if never recorded (force a full rebuild on next Excel open).
    """
    p = EXCEL_BUILD_STATE_PATH
    if not p.is_file():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    v = data.get("debug_mode")
    if not isinstance(v, str):
        return None
    s = v.strip().lower()
    if s in ("1", "true", "yes"):
        return "1"
    if s in ("0", "false", "no"):
        return "0"
    return None


def record_excel_build_debug_mode() -> None:
    """Persist :func:`current_excel_build_debug_flag` after a full rebuild or a refresh-only update."""
    p = EXCEL_BUILD_STATE_PATH
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(
            json.dumps({"debug_mode": current_excel_build_debug_flag()}, indent=2),
            encoding="utf-8",
        )
    except OSError:
        pass


def clean_value(value):
    """Return None for JSON null and for the literal string 'null'."""
    if value is None:
        return None
    if isinstance(value, str) and value.strip().lower() == "null":
        return None
    return value


def infer_company_from_subject(subject) -> str | None:
    subject = clean_value(subject)
    if not isinstance(subject, str):
        return None

    normalized = subject
    while True:
        updated = re.sub(r"^\s*(?:fw|fwd|re)\s*:\s*", "", normalized, flags=re.IGNORECASE)
        if updated == normalized:
            break
        normalized = updated

    patterns = [
        r"your\s+(.+?)\s+order(?:\b|:)",
        r"order\s+from\s+(.+?)(?:\b|:)",
        r"thanks\s+.+?\s+for\s+your\s+purchase\s+with\s+(.+?)(?:\b|!|\.|:)",
    ]

    for pattern in patterns:
        match = re.search(pattern, normalized, flags=re.IGNORECASE)
        if match:
            company = clean_value(match.group(1))
            if company:
                return company.strip(" -,:;.!?")

    return None


def _file_uri_to_os_path(uri: str) -> Path | None:
    try:
        parsed = urlparse(uri)
        if parsed.scheme != "file":
            return None
        path = url2pathname(unquote(parsed.path))
        return Path(path).resolve()
    except (ValueError, OSError):
        return None


def _html_file_uri_for_record(record: dict) -> str | None:
    """``email_contents/html/<pdf_stem>.html`` under :data:`PROJECT_ROOT` (``BASE_DIR``)."""
    pdf_uri = clean_value(record.get("source_file_link"))
    if not pdf_uri or not isinstance(pdf_uri, str) or not pdf_uri.startswith("file:///"):
        return None
    pdf_path = _file_uri_to_os_path(pdf_uri)
    if pdf_path is None:
        return None
    html_path = (
        PROJECT_ROOT / "email_contents" / "html" / f"{pdf_path.stem}.html"
    ).resolve()
    if not html_path.is_file():
        return None
    return "file:///" + str(html_path).replace("\\", "/")


def get_company_value(record: dict):
    explicit_company = clean_value(record.get("company"))
    if explicit_company:
        return explicit_company

    return infer_company_from_subject(record.get("subject"))


def _build_column_order() -> tuple[list[str], list[str]]:
    keys = _column_order_keys()
    labels = [COLUMN_HEADERS[k] for k in keys]
    return keys, labels


def _is_first_row_for_order(
    record: dict,
    batch_index: int,
    records: list[dict],
    *,
    sheet_row: int | None = None,
    ws=None,
    order_number_col_idx: int | None = None,
) -> bool:
    """True on the first Excel row for a given ``order_number`` (contiguous block)."""
    if batch_index > 0:
        return records[batch_index - 1].get("order_number") != record.get(
            "order_number"
        )
    if (
        sheet_row is not None
        and ws is not None
        and order_number_col_idx is not None
        and sheet_row > 2
    ):
        prev = ws.cell(row=sheet_row - 1, column=order_number_col_idx).value
        return prev != record.get("order_number")
    return True


def _record_to_row(
    record: dict,
    column_keys: list[str],
    *,
    shipping_status_first_row: bool = True,
) -> list:
    """Convert a single JSON record to a flat row list matching column_keys."""
    row: list = []
    for key in column_keys:
        if key == "company":
            row.append(get_company_value(record))
        elif key == "copy_file_path":
            row.append(clean_value(record.get("source_file_link")))
        elif key == "html_source_link":
            row.append(_html_file_uri_for_record(record))
        elif key == "gift_invoice_action":
            row.append(None)
        elif key == "open_tracking_numbers_order":
            row.append(None)
        elif key == "tracking_quick_status":
            row.append(
                display_shipping_status_for_record(
                    record,
                    shipping_status_first_row=shipping_status_first_row,
                )
            )
        else:
            row.append(clean_value(record.get(key)))
    return row


def style_header_row(ws, col_count: int, row_idx: int = HEADER_ROW):
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN


def _split_automation_hub_record(records: list[dict]) -> tuple[dict | None, list[dict]]:
    """Pull the workbook action row out of the data rows."""
    hub_record: dict | None = None
    data_records: list[dict] = []
    for record in records:
        if is_automation_hub_record(record):
            if hub_record is None:
                hub_record = record
            continue
        data_records.append(record)
    return hub_record, data_records


def style_automation_hub_row(ws, col_count: int, row_idx: int = ACTION_ROW) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = CATEGORY_FILLS[AUTOMATION_HUB_CATEGORY]
        cell.font = AUTOMATION_HUB_FONT
        cell.alignment = CENTER_ALIGN


def center_invoice_and_shipping_headers(ws, column_keys: list[str]) -> None:
    """Force center alignment for Invoice Link and Shipping Status (template/Excel defaults may differ)."""
    for key in ("gift_invoice_action", "tracking_quick_status"):
        for col_idx in _col_indices(column_keys, key):
            ws.cell(row=HEADER_ROW, column=col_idx).alignment = CENTER_ALIGN


def set_column_widths(ws, column_keys: list[str]):
    width_map = {
        "email_category":        22,
        "order_number":          18,
        "gift_invoice_action":   14,
        "purchase_datetime":     22,
        "company":               24,
        "email":                 30,
        "total_amount_paid":     14,
        "tax_paid":              12,
        "tracking_quick_status":       40,
        "open_tracking_list":           20,
        "open_tracking_numbers_web":  22,
        "open_tracking_numbers_order": 34,
        "copy_file_path":        12,
        "source_file_link":      12,
        "html_source_link":      12,
    }
    for col_idx, key in enumerate(column_keys, start=1):
        w = width_map.get(key, 16)
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _col_indices(column_keys: list[str], key: str) -> list[int]:
    """Return 1-based column indices for every occurrence of *key* in *column_keys*."""
    return [i + 1 for i, k in enumerate(column_keys) if k == key]


_FILE_LINK_DISPLAY = {
    "source_file_link": "View PDF",
    "html_source_link": "View HTML",
}


def apply_file_link_hyperlinks(ws, column_keys: list[str], start_row: int) -> None:
    """Replace file URI values with clickable hyperlinks (View PDF / View HTML)."""
    for key in ("source_file_link", "html_source_link"):
        if key not in column_keys:
            continue
        label = _FILE_LINK_DISPLAY[key]
        for col_idx in _col_indices(column_keys, key):
            for row in ws.iter_rows(
                min_row=start_row,
                max_row=ws.max_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                cell = row[0]
                uri = cell.value
                if uri and isinstance(uri, str) and uri.startswith("file:///"):
                    cell.value = label
                    cell.hyperlink = uri
                    cell.font = HYPERLINK_FONT
                    cell.alignment = CENTER_ALIGN


def _sheet_name_for_excel_ref(name: str) -> str:
    """Build sheet name token for #'Name'!$A$1 style hyperlinks."""
    if "'" in name:
        return "'" + name.replace("'", "''") + "'"
    if " " in name or not name.replace("_", "").isalnum():
        return "'" + name + "'"
    return name


def _set_internal_hyperlink(cell, location: str) -> None:
    """Set a true in-workbook hyperlink, not an external target beginning with ``#``."""
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=location)


def apply_copy_path_hyperlink_columns(
    ws,
    col_indices: list[int],
    start_row: int,
    records: list[dict],
    *,
    anchor_col_idx: int | None = None,
) -> None:
    """Open File Location: in-workbook # links only; real file URI in hidden column COPY_PATH_URI_COL."""
    col_uri = COPY_PATH_URI_COL
    sn = _sheet_name_for_excel_ref(ws.title)
    ws.column_dimensions[get_column_letter(col_uri)].hidden = True
    if not records:
        return

    for i, record in enumerate(records):
        row_idx = start_row + i
        uri = clean_value(record.get("source_file_link"))
        if uri and isinstance(uri, str) and uri.startswith("file:///"):
            ws.cell(row=row_idx, column=col_uri, value=uri)

    for col_idx in col_indices:
        for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(records) - 1,
                                min_col=col_idx, max_col=col_idx):
            cell = row[0]
            row_num = cell.row
            stored = ws.cell(row=row_num, column=col_uri).value
            cell.value = None
            cell.hyperlink = None
            if not stored or not isinstance(stored, str) or not stored.startswith("file:///"):
                continue
            cell.value = "File Loc"
            anchor_letter = get_column_letter(anchor_col_idx or col_idx)
            _set_internal_hyperlink(cell, f"{sn}!${anchor_letter}${row_num}")
            cell.font = HYPERLINK_FONT
            cell.alignment = CENTER_ALIGN


def _tracking_urls_for_record(record: dict) -> list[str]:
    """URLs from JSON ``tracking_links`` for hidden columns / View tracking links."""
    raw = record.get("tracking_links")
    if not isinstance(raw, list) or not raw:
        return []
    out: list[str] = []
    for u in raw:
        if not isinstance(u, str):
            continue
        s = u.strip()
        if not s or s == MULTIPLE_TRACKING_LINKS:
            continue
        out.append(s)
    return out


def resolve_shipping_summary(record: dict) -> str | None:
    """Aggregate delivered fraction from 17TRACK cache (see prefetch before Excel build)."""
    nums = _tracking_numbers_for_record(record)
    if not nums:
        return None
    try:
        from trackingNumbersViewer.seventeen_track_smart import (
            format_shipping_summary_line,
            shipping_summary_metrics,
        )

        valid, delivered = shipping_summary_metrics(nums)
        return format_shipping_summary_line(valid, delivered)
    except Exception:
        return None


def display_shipping_status_for_record(
    record: dict,
    *,
    shipping_status_first_row: bool = True,
) -> str | None:
    if is_automation_hub_record(record):
        return AUTOMATION_HUB_STATUS_LABEL
    if is_pod_record(record):
        return None
    if not shipping_status_first_row:
        return None
    return resolve_shipping_summary(record)


def _shipping_summary_color_percent(text) -> int | None:
    """Map cell text to 0–100 for conditional font color; ``None`` = leave default."""
    if text is None:
        return None
    if not isinstance(text, str):
        return None
    s = text.strip()
    if not s:
        return None
    if s == "All Delivered":
        return 100
    if s == "None Delivered":
        return 0
    m = re.match(r"^(\d{1,3})%\s*Delivered\s*$", s, re.I)
    if m:
        return max(0, min(100, int(m.group(1))))
    return None


def apply_shipping_summary_cells(
    ws,
    column_keys: list[str],
    records: list[dict],
    start_row: int,
    *,
    vba_friendly: bool = False,
    end_row: int | None = None,
    shipping_status_col_idx: int | None = None,
    order_number_col_idx: int | None = None,
) -> None:
    """Shipping Status text with % colors; in .xlsm, same cell links to the status viewer (VBA)."""
    if shipping_status_col_idx is not None:
        col_idx = shipping_status_col_idx
    else:
        cols = _col_indices(column_keys, "tracking_quick_status")
        if not cols:
            return
        col_idx = cols[0]
    if order_number_col_idx is None:
        order_number_col_idx = column_keys.index("order_number") + 1
    last_row = end_row if end_row is not None else ws.max_row
    red = Font(name="Calibri", color="C00000")
    yellow = Font(name="Calibri", color="BF9000")
    green = Font(name="Calibri", color="006100")
    red_u = Font(name="Calibri", color="C00000", underline="single")
    yellow_u = Font(name="Calibri", color="BF9000", underline="single")
    green_u = Font(name="Calibri", color="006100", underline="single")
    hub_u = AUTOMATION_HUB_LINK_FONT
    hub_plain = AUTOMATION_HUB_FONT
    sn = _sheet_name_for_excel_ref(ws.title)
    cat_col = column_keys.index("email_category") + 1
    anchor_letter = get_column_letter(cat_col)

    for offset, row_idx in enumerate(range(start_row, last_row + 1)):
        record = records[offset] if offset < len(records) else {}
        cell = ws.cell(row=row_idx, column=col_idx)
        is_first = _is_first_row_for_order(
            record,
            offset,
            records,
            sheet_row=row_idx,
            ws=ws,
            order_number_col_idx=order_number_col_idx,
        )
        if not is_first:
            cell.value = None
            cell.hyperlink = None
            cell.font = CELL_FONT
            cell.alignment = CENTER_ALIGN
            continue

        nums = _tracking_numbers_for_record(record)
        raw = cell.value
        if isinstance(raw, str):
            display = raw.strip() or None
        else:
            display = raw if raw not in (None, "") else None
        if is_automation_hub_record(record):
            if display:
                cell.value = display
            if vba_friendly and display:
                _set_internal_hyperlink(cell, f"{sn}!${anchor_letter}${row_idx}")
                cell.font = hub_u
            else:
                cell.hyperlink = None
                cell.font = hub_plain
            cell.alignment = CENTER_ALIGN
            continue
        if is_pod_record(record):
            cell.value = None
            cell.hyperlink = None
            cell.font = CELL_FONT
            cell.alignment = CENTER_ALIGN
            continue
        if nums and not display:
            cell.value = "View Shipping Status"
            display = cell.value
        pct = _shipping_summary_color_percent(display)

        if nums and vba_friendly and display:
            _set_internal_hyperlink(cell, f"{sn}!${anchor_letter}${row_idx}")
            if pct is not None:
                if pct < 33:
                    cell.font = red_u
                elif pct <= 66:
                    cell.font = yellow_u
                else:
                    cell.font = green_u
            else:
                cell.font = HYPERLINK_FONT
            cell.alignment = CENTER_ALIGN
            continue

        cell.hyperlink = None
        if pct is not None:
            if pct < 33:
                cell.font = red
            elif pct <= 66:
                cell.font = yellow
            else:
                cell.font = green
        else:
            cell.font = CELL_FONT
        cell.alignment = CENTER_ALIGN


def _tracking_numbers_for_record(record: dict) -> list[str]:
    """Plain tracking IDs from JSON ``tracking_numbers`` for hidden columns / viewers."""
    raw = record.get("tracking_numbers")
    if not isinstance(raw, list) or not raw:
        return []
    out: list[str] = []
    seen: set[str] = set()
    for x in raw:
        if not isinstance(x, str):
            continue
        s = x.strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def _aggregate_tracking_data_for_order_block(
    records: list[dict], start_idx: int
) -> tuple[dict[str, int], dict[str, bool]]:
    """Count each tracking ID across contiguous rows with the same ``order_number``; link = any row confirms."""
    from collections import Counter

    counts: Counter[str] = Counter()
    link_any: dict[str, bool] = {}
    if start_idx >= len(records):
        return {}, {}
    order_key = records[start_idx].get("order_number")
    j = start_idx
    while j < len(records) and records[j].get("order_number") == order_key:
        rec = records[j]
        nums = _tracking_numbers_for_record(rec)
        raw_flags = rec.get("tracking_numbers_link_confirmed")
        for idx, num in enumerate(nums):
            counts[num] += 1
            ok = (
                bool(raw_flags[idx])
                if isinstance(raw_flags, list) and idx < len(raw_flags)
                else False
            )
            link_any[num] = link_any.get(num, False) or ok
        j += 1
    return dict(counts), link_any


def _tracking_number_confirm_flags_for_record(record: dict) -> list[str]:
    """Parallel ``1``/``\"\"`` flags for hidden confirm columns (aligned to ``tracking_numbers`` slots)."""
    nums = _tracking_numbers_for_record(record)
    raw = record.get("tracking_numbers_link_confirmed")
    flags: list[bool] = []
    if isinstance(raw, list):
        for i in range(len(nums)):
            flags.append(bool(raw[i]) if i < len(raw) else False)
    else:
        flags = [False] * len(nums)
    out: list[str] = []
    for i in range(TRACKING_NUMBER_SLOT_COUNT):
        if i < len(flags) and flags[i]:
            out.append("1")
        else:
            out.append("")
    return out


def apply_hidden_tracking_url_columns(
    ws,
    records: list[dict],
    start_row: int,
    *,
    vba_friendly: bool = False,
):
    """Write JSON ``tracking_links`` into hidden columns for VBA / View tracking links viewer."""
    if not records:
        return
    last_uri_col = TRACKING_URI_COL_START + TRACKING_URI_SLOT_COUNT - 1

    for i, record in enumerate(records):
        row_idx = start_row + i
        urls = _tracking_urls_for_record(record)
        for uc in range(TRACKING_URI_COL_START, last_uri_col + 1):
            ws.cell(row=row_idx, column=uc, value=None)
        if not vba_friendly:
            continue
        for j, u in enumerate(urls[:TRACKING_URI_SLOT_COUNT]):
            ws.cell(row=row_idx, column=TRACKING_URI_COL_START + j, value=u)

    if vba_friendly:
        for uc in range(TRACKING_URI_COL_START, last_uri_col + 1):
            ws.column_dimensions[get_column_letter(uc)].hidden = True


def apply_hidden_tracking_number_columns(
    ws,
    records: list[dict],
    start_row: int,
    *,
    vba_friendly: bool = False,
):
    """Write JSON ``tracking_numbers`` + link-cross-check flags into hidden columns for VBA."""
    if not records:
        return
    last_num = TRACKING_NUMBER_COL_START + TRACKING_NUMBER_SLOT_COUNT - 1
    last_conf = TRACKING_NUM_CONFIRM_COL_START + TRACKING_NUM_CONFIRM_SLOT_COUNT - 1

    for i, record in enumerate(records):
        row_idx = start_row + i
        nums = _tracking_numbers_for_record(record)
        confirms = _tracking_number_confirm_flags_for_record(record)
        for uc in range(TRACKING_NUMBER_COL_START, last_num + 1):
            ws.cell(row=row_idx, column=uc, value=None)
        for uc in range(TRACKING_NUM_CONFIRM_COL_START, last_conf + 1):
            ws.cell(row=row_idx, column=uc, value=None)
        if not vba_friendly:
            continue
        for j, n in enumerate(nums[:TRACKING_NUMBER_SLOT_COUNT]):
            ws.cell(row=row_idx, column=TRACKING_NUMBER_COL_START + j, value=n)
        for j, fl in enumerate(confirms[:TRACKING_NUM_CONFIRM_SLOT_COUNT]):
            ws.cell(row=row_idx, column=TRACKING_NUM_CONFIRM_COL_START + j, value=fl or None)

    if vba_friendly:
        for uc in range(TRACKING_NUMBER_COL_START, last_num + 1):
            ws.column_dimensions[get_column_letter(uc)].hidden = True
        for uc in range(TRACKING_NUM_CONFIRM_COL_START, last_conf + 1):
            ws.column_dimensions[get_column_letter(uc)].hidden = True


def apply_gift_invoice_link_columns(
    ws,
    column_keys: list[str],
    records: list[dict],
    start_row: int,
    project_root: Path,
) -> None:
    """Fill ``Invoice link`` from ``gift_invoice_links.json`` (order-number–based edges)."""
    gift_cols = _col_indices(column_keys, "gift_invoice_action")
    if not gift_cols or not records:
        return
    gcol = gift_cols[0]
    sn = _sheet_name_for_excel_ref(ws.title)
    # Anchor on Category (same row), not the Invoice link cell. Self-referential #links
    # often skip Workbook_SheetFollowHyperlink because Excel does not change selection.
    cat_col = column_keys.index("email_category") + 1
    anchor_letter = get_column_letter(cat_col)
    link_path = links_path_for_project_root(project_root)
    edges = load_edges(link_path, records)

    for i, record in enumerate(records):
        row_idx = start_row + i
        key = stable_record_key(record, i)
        cat = record.get("email_category")
        ordn = normalized_order_number(record)
        label = gift_order_link_label(
            cat if isinstance(cat, str) else None,
            key,
            ordn,
            edges,
        )
        gc = ws.cell(row=row_idx, column=gcol)
        gc.hyperlink = None
        if label:
            gc.value = label
            _set_internal_hyperlink(gc, f"{sn}!${anchor_letter}${row_idx}")
            gc.font = HYPERLINK_FONT
            gc.alignment = CENTER_ALIGN
            src = ws.cell(row=row_idx, column=cat_col)
            # StyleProxy from the template is not hashable for openpyxl's style table; copy unwraps it.
            gc.fill = copy(src.fill)
        else:
            gc.value = None


def apply_open_tracking_list_column(
    ws,
    column_keys: list[str],
    records: list[dict],
    start_row: int,
    *,
    vba_friendly: bool = False,
):
    """``View tracking links`` cell: VBA launches :mod:`trackingLinkViewer` with all row tracking URLs."""
    cols = _col_indices(column_keys, "open_tracking_list")
    if not cols or not records:
        return
    col_idx = cols[0]
    sn = _sheet_name_for_excel_ref(ws.title)
    anchor_letter = get_column_letter(column_keys.index("email_category") + 1)

    for i, record in enumerate(records):
        row_idx = start_row + i
        cell = ws.cell(row=row_idx, column=col_idx)
        urls = _tracking_urls_for_record(record)
        cell.hyperlink = None
        if not urls:
            cell.value = None
            continue
        if vba_friendly:
            cell.value = COLUMN_HEADERS["open_tracking_list"]
            _set_internal_hyperlink(cell, f"{sn}!${anchor_letter}${row_idx}")
            cell.font = HYPERLINK_FONT
            cell.alignment = CENTER_ALIGN
        else:
            cell.value = "—"
            cell.alignment = CENTER_ALIGN


def apply_open_tracking_numbers_web_column(
    ws,
    column_keys: list[str],
    records: list[dict],
    start_row: int,
    *,
    vba_friendly: bool = False,
):
    """``View Tracking Numbers``: VBA launches :mod:`tracking_numbers_viewer` in ``web`` mode."""
    cols = _col_indices(column_keys, "open_tracking_numbers_web")
    if not cols or not records:
        return
    col_idx = cols[0]
    sn = _sheet_name_for_excel_ref(ws.title)
    anchor_letter = get_column_letter(column_keys.index("email_category") + 1)

    for i, record in enumerate(records):
        row_idx = start_row + i
        cell = ws.cell(row=row_idx, column=col_idx)
        nums = _tracking_numbers_for_record(record)
        cell.hyperlink = None
        if not nums:
            cell.value = None
            continue
        if vba_friendly:
            cell.value = COLUMN_HEADERS["open_tracking_numbers_web"]
            _set_internal_hyperlink(cell, f"{sn}!${anchor_letter}${row_idx}")
            cell.font = HYPERLINK_FONT
            cell.alignment = CENTER_ALIGN
        else:
            cell.value = "—"
            cell.alignment = CENTER_ALIGN


def apply_open_tracking_numbers_order_column(
    ws,
    column_keys: list[str],
    records: list[dict],
    start_row: int,
    *,
    vba_friendly: bool = False,
):
    """First row per order only: union of tracking IDs with counts; VBA launches ``order`` viewer mode."""
    cols = _col_indices(column_keys, "open_tracking_numbers_order")
    if not cols or not records:
        return
    col_idx = cols[0]
    sn = _sheet_name_for_excel_ref(ws.title)
    anchor_letter = get_column_letter(column_keys.index("email_category") + 1)
    label = COLUMN_HEADERS["open_tracking_numbers_order"]

    for i, record in enumerate(records):
        row_idx = start_row + i
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.hyperlink = None
        if not _is_first_row_for_order(record, i, records):
            cell.value = None
            continue
        counts, _link_any = _aggregate_tracking_data_for_order_block(records, i)
        if not counts:
            cell.value = None
            continue
        if vba_friendly:
            cell.value = label
            _set_internal_hyperlink(cell, f"{sn}!${anchor_letter}${row_idx}")
            cell.font = HYPERLINK_FONT
            cell.alignment = CENTER_ALIGN
        else:
            cell.value = "—"
            cell.alignment = CENTER_ALIGN


def apply_debug_gated_tracking_tool_columns(ws, column_keys: list[str]) -> None:
    """Hide viewer link columns unless ``DEBUG_MODE`` is enabled (see :func:`runLogger.is_debug`)."""
    if is_debug():
        return
    for key in (
        "open_tracking_list",
        "open_tracking_numbers_web",
        "open_tracking_numbers_order",
    ):
        for col_idx in _col_indices(column_keys, key):
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True


def _merge_border(cell, top=None, bottom=None, left=None, right=None):
    """Merge new sides into a cell's existing border without overwriting unset sides."""
    old = cell.border
    cell.border = Border(
        top=top if top is not None else old.top,
        bottom=bottom if bottom is not None else old.bottom,
        left=left if left is not None else old.left,
        right=right if right is not None else old.right,
    )


def apply_cell_styles(ws, start_row: int):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = CELL_FONT
            cell.alignment = LEFT_ALIGN


def apply_category_colors(ws, start_row: int, category_col: int):
    """Shade every cell in a row with a pastel fill based on its email_category."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        category = ws.cell(row=row[0].row, column=category_col).value
        fill = CATEGORY_FILLS.get(category)
        if fill:
            for cell in row:
                cell.fill = fill


def _order_band_by_row(records: list[dict]) -> list[int]:
    """0/1 alternating band per distinct order_number (first in file = 0); -1 = no order id."""
    seen: dict[str, int] = {}
    next_band_slot = 0
    out: list[int] = []
    for r in records:
        on = normalized_order_number(r)
        if not on:
            out.append(-1)
            continue
        if on not in seen:
            seen[on] = next_band_slot % 2
            next_band_slot += 1
        out.append(seen[on])
    return out


def apply_order_number_row_colors(ws, start_row: int, records: list[dict]) -> None:
    """Green/blue bands by order_number; Gift Card stays yellow; missing order_number → purple."""
    bands = _order_band_by_row(records)
    for offset, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        rec = records[offset] if offset < len(records) else {}
        band = bands[offset] if offset < len(bands) else -1
        cat = rec.get("email_category")
        if cat == AUTOMATION_HUB_CATEGORY:
            fill = CATEGORY_FILLS[AUTOMATION_HUB_CATEGORY]
        elif cat == "Gift Card":
            fill = CATEGORY_FILLS["Gift Card"]
        elif band < 0:
            fill = UNKNOWN_ORDER_FILL
        else:
            fill = ORDER_BAND_FILLS[band]
        for cell in row:
            cell.fill = fill


def apply_special_row_styles(ws, start_row: int, records: list[dict], column_keys: list[str]) -> None:
    shipping_col = column_keys.index("tracking_quick_status") + 1
    for offset, row_idx in enumerate(range(start_row, ws.max_row + 1)):
        rec = records[offset] if offset < len(records) else {}
        if not is_automation_hub_record(rec):
            continue
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = AUTOMATION_HUB_FONT
            cell.alignment = CENTER_ALIGN if col_idx == shipping_col else LEFT_ALIGN
        hub_cell = ws.cell(row=row_idx, column=shipping_col)
        if hub_cell.value:
            hub_cell.font = AUTOMATION_HUB_LINK_FONT if hub_cell.hyperlink else AUTOMATION_HUB_FONT
            hub_cell.alignment = CENTER_ALIGN


def apply_row_borders(ws, start_row: int):
    """Add a hairline bottom border on every row from the header through the last data row."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            _merge_border(cell, bottom=HAIR_SIDE)


def apply_section_dividers(ws, boundary_cols: list[int]):
    """Draw a hairline right border on each listed column (all rows)."""
    for col_idx in boundary_cols:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            _merge_border(row[0], right=HAIR_SIDE)


def apply_table_outline(ws):
    """Draw a hairline border around the entire table (action row + header + data)."""
    last_row = ws.max_row
    last_col = ws.max_column

    for col_idx in range(1, last_col + 1):
        _merge_border(ws.cell(row=ACTION_ROW, column=col_idx), top=HAIR_SIDE)
        _merge_border(ws.cell(row=last_row, column=col_idx), bottom=HAIR_SIDE)

    for row_idx in range(ACTION_ROW, last_row + 1):
        _merge_border(ws.cell(row=row_idx, column=1), left=HAIR_SIDE)
        _merge_border(ws.cell(row=row_idx, column=last_col), right=HAIR_SIDE)


def apply_header_border(ws):
    """Draw a hairline top and bottom border on the header row (horizontal only)."""
    last_col = ws.max_column
    for col_idx in range(1, last_col + 1):
        _merge_border(ws.cell(row=HEADER_ROW, column=col_idx),
                      top=HAIR_SIDE, bottom=HAIR_SIDE)


def populate_orders_sheet(wb: Workbook, records: list[dict]) -> None:
    """Rebuild the Orders sheet from scratch (headers + rows + styling + hyperlinks)."""
    vba_friendly = getattr(wb, "vba_archive", None) is not None
    column_keys, header_labels = _build_column_order()
    hub_record, data_records = _split_automation_hub_record(records)

    if "Orders" in wb.sheetnames:
        ws = wb["Orders"]
    else:
        ws = wb.active
        ws.title = "Orders"

    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    if hub_record is None:
        hub_record = automation_hub_record()
    ws.append(_record_to_row(hub_record, column_keys, shipping_status_first_row=True))
    ws.append(header_labels)
    style_header_row(ws, len(column_keys))
    style_automation_hub_row(ws, len(column_keys))

    for i, record in enumerate(data_records):
        first = _is_first_row_for_order(record, i, data_records)
        ws.append(
            _record_to_row(record, column_keys, shipping_status_first_row=first)
        )

    copy_cols = _col_indices(column_keys, "copy_file_path")
    section_boundary_cols = [
        idx
        for key in _section_boundary_keys(column_keys)
        for idx in _col_indices(column_keys, key)
    ]

    apply_cell_styles(ws, start_row=DATA_START_ROW)
    set_column_widths(ws, column_keys)

    category_col_idx = column_keys.index("email_category") + 1
    if _use_legacy_category_row_colors():
        apply_category_colors(ws, start_row=DATA_START_ROW, category_col=category_col_idx)
    else:
        apply_order_number_row_colors(ws, start_row=DATA_START_ROW, records=data_records)
    apply_shipping_summary_cells(
        ws, column_keys, data_records, start_row=DATA_START_ROW, vba_friendly=vba_friendly
    )
    apply_shipping_summary_cells(
        ws,
        column_keys,
        [hub_record],
        start_row=ACTION_ROW,
        vba_friendly=vba_friendly,
        end_row=ACTION_ROW,
    )

    apply_row_borders(ws, start_row=DATA_START_ROW)
    apply_header_border(ws)
    apply_section_dividers(ws, section_boundary_cols)
    apply_table_outline(ws)

    apply_copy_path_hyperlink_columns(
        ws,
        copy_cols,
        start_row=DATA_START_ROW,
        records=data_records,
        anchor_col_idx=category_col_idx,
    )
    apply_file_link_hyperlinks(ws, column_keys, start_row=DATA_START_ROW)
    apply_hidden_tracking_url_columns(ws, data_records, start_row=DATA_START_ROW, vba_friendly=vba_friendly)
    apply_hidden_tracking_number_columns(ws, data_records, start_row=DATA_START_ROW, vba_friendly=vba_friendly)
    apply_open_tracking_list_column(
        ws, column_keys, data_records, start_row=DATA_START_ROW, vba_friendly=vba_friendly
    )
    apply_open_tracking_numbers_web_column(
        ws, column_keys, data_records, start_row=DATA_START_ROW, vba_friendly=vba_friendly,
    )
    apply_open_tracking_numbers_order_column(
        ws, column_keys, data_records, start_row=DATA_START_ROW, vba_friendly=vba_friendly,
    )
    apply_gift_invoice_link_columns(ws, column_keys, data_records, DATA_START_ROW, PROJECT_ROOT)
    apply_debug_gated_tracking_tool_columns(ws, column_keys)
    apply_special_row_styles(ws, start_row=ACTION_ROW, records=[hub_record], column_keys=column_keys)
    apply_special_row_styles(ws, start_row=DATA_START_ROW, records=data_records, column_keys=column_keys)

    center_invoice_and_shipping_headers(ws, column_keys)

    ws.freeze_panes = FREEZE_PANES_CELL


def _emit_excel_launcher_progress(pct: int, msg: str = "") -> None:
    """One-line progress for email_sorter_launcher when EXCEL_LAUNCHER_PROGRESS=1."""
    if os.getenv("EXCEL_LAUNCHER_PROGRESS") != "1":
        return
    pct = max(0, min(100, int(pct)))
    line = f"EMAIL_SORTER_EXCEL_PROGRESS pct={pct}"
    if msg:
        safe = msg.replace("\n", " ").replace("\r", "")[:140]
        line += f" msg={safe}"
    print(line, flush=True)


def _excel_launcher_17track_skip_requested() -> bool:
    """True when launcher set ``EMAIL_SORTER_17TRACK_SKIP_FLAG`` (user chose Skip 17Track)."""
    raw = (os.getenv("EMAIL_SORTER_17TRACK_SKIP_FLAG") or "").strip()
    if not raw:
        return False
    try:
        p = Path(raw)
        return p.is_file() and p.read_text(encoding="utf-8").strip() == "1"
    except OSError:
        return False


def set_clipboard_ini_cell(wb: Workbook, ini_path: Path) -> None:
    """Write absolute ini path for Workbook_SheetFollowHyperlink; hide column AA."""
    ws = wb["Orders"]
    ws[CLIPBOARD_INI_CELL] = str(ini_path.resolve())
    ws.column_dimensions["AA"].hidden = True


def _prefetch_17track_for_excel_build(
    records: list[dict],
    *,
    quiet: bool = False,
    on_prefetch_progress: Callable[[int, int], None] | None = None,
    cancel_check: Callable[[], bool] | None = None,
) -> None:
    """Prefetch 17TRACK rows for *records*.

    When env ``EMAIL_SORTER_17TRACK_QUOTA_SESSION=1`` (Email Sorter launcher **Run** or **Excel**),
    performs exactly **two** ``getquota`` checks with user dialogs: once before prefetch and
    once after prefetch completes. Other callers (no env var) perform **no** quota checks here.
    """
    session = (os.getenv("EMAIL_SORTER_17TRACK_QUOTA_SESSION") or "").strip() == "1"
    try:
        from trackingNumbersViewer.seventeen_track_api import api_key_from_env
        from trackingNumbersViewer.seventeen_track_smart import prefetch_tracking_for_records
        from shared.load_17track_quota import get_17track_quota_module

        if session and api_key_from_env():
            _, skip = get_17track_quota_module().quota_prefetch_gate()
            if skip:
                try:
                    from shared import runLogger as RL

                    RL.log(
                        "17track",
                        f"{RL.ts()} prefetch skipped after quota gate (0 remaining)",
                    )
                except Exception:
                    pass
                print(
                    "WARNING: 17TRACK prefetch skipped — API quota is exhausted (0 remaining).",
                    file=sys.stderr,
                )
                return

        if not quiet:
            print("[createExcelDocument] Prefetching 17TRACK data for tracking numbers …")
        prefetch_tracking_for_records(
            records,
            on_progress=on_prefetch_progress,
            cancel_check=cancel_check,
        )
    except Exception as e:
        if not quiet:
            print(f"[createExcelDocument] Tracking prefetch skipped: {e}")
    finally:
        if session:
            try:
                from trackingNumbersViewer.seventeen_track_api import api_key_from_env
                from shared.load_17track_quota import get_17track_quota_module

                if api_key_from_env():
                    get_17track_quota_module().quota_session_end_notify()
            except Exception:
                pass


def _build_populated_orders_workbook(
    records: list[dict], *, verbose_clipboard_log: bool = False
) -> tuple[Workbook, bool]:
    """Load macro template (or a plain workbook), populate Orders, set AA1 ini when .xlsm."""
    using_template = ORDERS_TEMPLATE_PATH.is_file()
    macro_mod = None
    if sys.platform == "win32":
        macro_mod = _macro_template_module()
    if macro_mod is not None:
        refreshed = False
        try:
            refreshed = bool(macro_mod.refresh_macro_template(ORDERS_TEMPLATE_PATH))
        except Exception as exc:
            print(f"[createExcelDocument] Macro template refresh skipped: {exc}")
        if refreshed:
            using_template = ORDERS_TEMPLATE_PATH.is_file()
        elif not using_template:
            if macro_mod.ensure_macro_template(ORDERS_TEMPLATE_PATH):
                using_template = ORDERS_TEMPLATE_PATH.is_file()

    if using_template:
        wb = load_workbook(str(ORDERS_TEMPLATE_PATH), keep_vba=True)
        populate_orders_sheet(wb, records)
    else:
        wb = build_workbook(records)

    if using_template:
        m = macro_mod if macro_mod is not None else _macro_template_module()
        script_path = Path(__file__).resolve().parent / "copy_email_path_to_clipboard.py"
        ini_written = m.write_clipboard_launch_ini(
            CLIPBOARD_LAUNCH_INI_PATH,
            sys.executable,
            script_path,
            viewer_script=TRACKING_VIEWER_SCRIPT,
            giftcard_link_script=GIFT_INVOICE_LINK_SCRIPT,
            tracking_numbers_viewer_script=TRACKING_NUMBERS_VIEWER_SCRIPT,
            tracking_status_viewer_script=TRACKING_STATUS_VIEWER_SCRIPT,
            pod_workflow_script=POD_WORKFLOW_SCRIPT,
        )
        set_clipboard_ini_cell(wb, ini_written)
        if verbose_clipboard_log:
            print(f"Wrote clipboard launcher config: {ini_written}")

    return wb, using_template


def rebuild_orders_workbook(excel_output_path: str | Path) -> None:
    """
    Full rebuild from ``results.json`` (same as a CLI run, minus duplicate-flag reset).

    Column layout — including whether **View HTML** and the tracking tool columns are visible —
    follows **current** ``DEBUG_MODE`` (see :func:`runLogger.is_debug`). Call
    ``apply_runtime_settings_from_json()`` before loading this module if settings may have changed.
    """
    _emit_excel_launcher_progress(2, "Loading results.json")
    records = load_excel_records(PROJECT_ROOT, include_automation_hub=True, sync_pod_json=True)

    def _on_track(done: int, total: int) -> None:
        if total <= 0:
            _emit_excel_launcher_progress(50, "17TRACK (no IDs)")
            return
        # Reserve 5–95% for 17TRACK; workbook build uses the tail.
        pct = 5 + int(90 * done / max(total, 1))
        _emit_excel_launcher_progress(pct, f"17TRACK {done}/{total}")

    _emit_excel_launcher_progress(4, "Prefetching 17TRACK…")
    _prefetch_17track_for_excel_build(
        records,
        quiet=True,
        on_prefetch_progress=_on_track,
        cancel_check=_excel_launcher_17track_skip_requested,
    )
    _emit_excel_launcher_progress(96, "Building workbook…")
    wb, _using_template = _build_populated_orders_workbook(records, verbose_clipboard_log=False)
    _emit_excel_launcher_progress(98, "Saving…")
    out = Path(excel_output_path).expanduser().resolve()
    wb.save(str(out))
    _emit_excel_launcher_progress(100, "Done")
    record_excel_build_debug_mode()


def build_workbook(records: list[dict]) -> Workbook:
    wb = Workbook()
    populate_orders_sheet(wb, records)
    return wb


def _load_workbook_editable(path: str) -> Workbook:
    """Preserve VBA when editing macro-enabled workbooks."""
    if Path(path).suffix.lower() == ".xlsm":
        return load_workbook(path, keep_vba=True)
    return load_workbook(path)


def _header_cell_matches_shipping_status(val) -> bool:
    if not isinstance(val, str):
        return False
    return val.strip() in SHIPPING_STATUS_HEADER_ALIASES


def refresh_orders_workbook_shipping_status(excel_path: str | Path) -> None:
    """
    Run smart 17TRACK prefetch (same rules as Excel build — skip needless terminal refreshes),
    then rewrite the Shipping Status column from ``results.json`` (one cell per order block).
    """
    path = Path(excel_path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(str(path))
    records = load_excel_records(PROJECT_ROOT, include_automation_hub=True, sync_pod_json=True)
    _hub_record, data_records = _split_automation_hub_record(records)
    try:
        _prefetch_17track_for_excel_build(records, quiet=True)
    except Exception:
        pass

    wb = _load_workbook_editable(str(path))
    ws = wb["Orders"] if "Orders" in wb.sheetnames else wb.active
    column_keys, _ = _build_column_order()

    shipping_col_idx: int | None = None
    for c in range(1, ws.max_column + 1):
        if _header_cell_matches_shipping_status(ws.cell(row=HEADER_ROW, column=c).value):
            shipping_col_idx = c
            break
    if shipping_col_idx is None:
        raise ValueError(
            "Could not find the Shipping Status column in the header row. Regenerate the workbook."
        )
    hdr = ws.cell(row=HEADER_ROW, column=shipping_col_idx)
    if hdr.value != COLUMN_HEADERS["tracking_quick_status"]:
        hdr.value = COLUMN_HEADERS["tracking_quick_status"]
    center_invoice_and_shipping_headers(ws, column_keys)

    order_col_idx: int | None = None
    want_order = COLUMN_HEADERS["order_number"]
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str) and v.strip() == want_order:
            order_col_idx = c
            break
    if order_col_idx is None:
        order_col_idx = column_keys.index("order_number") + 1

    n_rows = min(ws.max_row - DATA_START_ROW + 1, len(data_records))
    if n_rows < 1:
        wb.save(str(path))
        record_excel_build_debug_mode()
        return

    for offset in range(n_rows):
        row_idx = DATA_START_ROW + offset
        record = data_records[offset]
        cell = ws.cell(row=row_idx, column=shipping_col_idx)
        is_first = _is_first_row_for_order(
            record,
            offset,
            data_records,
            sheet_row=row_idx,
            ws=ws,
            order_number_col_idx=order_col_idx,
        )
        if not is_first:
            cell.value = None
            continue
        cell.value = display_shipping_status_for_record(
            record,
            shipping_status_first_row=is_first,
        )

    vba_friendly = path.suffix.lower() == ".xlsm" and getattr(
        wb, "vba_archive", None
    ) is not None
    apply_shipping_summary_cells(
        ws,
        column_keys,
        data_records[:n_rows],
        start_row=DATA_START_ROW,
        vba_friendly=vba_friendly,
        end_row=DATA_START_ROW + n_rows - 1,
        shipping_status_col_idx=shipping_col_idx,
        order_number_col_idx=order_col_idx,
    )
    wb.save(str(path))
    record_excel_build_debug_mode()


def append_to_workbook(path: str, records: list[dict]):
    try:
        _prefetch_17track_for_excel_build(records, quiet=True)
    except Exception:
        pass

    wb = _load_workbook_editable(path)
    ws = wb.active
    vba_friendly = Path(path).suffix.lower() == ".xlsm" and getattr(
        wb, "vba_archive", None
    ) is not None

    existing_headers = [
        ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)
    ]

    desired_keys, desired_labels = _build_column_order()

    if desired_labels != existing_headers:
        for col_idx, label in enumerate(desired_labels, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
        for col_idx in range(len(desired_labels) + 1, ws.max_column + 1):
            ws.cell(row=HEADER_ROW, column=col_idx, value=None)
        set_column_widths(ws, desired_keys)

    next_row = ws.max_row + 1
    order_col_idx = desired_keys.index("order_number") + 1

    for offset, record in enumerate(records):
        sheet_row = next_row + offset
        first = _is_first_row_for_order(
            record,
            offset,
            records,
            sheet_row=sheet_row,
            ws=ws,
            order_number_col_idx=order_col_idx,
        )
        row_data = _record_to_row(
            record, desired_keys, shipping_status_first_row=first
        )
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row + offset, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = LEFT_ALIGN

    copy_cols = _col_indices(desired_keys, "copy_file_path")
    section_boundary_cols = [
        idx
        for key in _section_boundary_keys(desired_keys)
        for idx in _col_indices(desired_keys, key)
    ]

    category_col_idx = desired_keys.index("email_category") + 1
    if _use_legacy_category_row_colors():
        apply_category_colors(ws, start_row=next_row, category_col=category_col_idx)
    else:
        apply_order_number_row_colors(ws, start_row=next_row, records=records)
    apply_shipping_summary_cells(
        ws,
        desired_keys,
        records,
        start_row=next_row,
        vba_friendly=vba_friendly,
        end_row=next_row + len(records) - 1,
    )

    apply_row_borders(ws, start_row=next_row)
    apply_header_border(ws)
    apply_section_dividers(ws, section_boundary_cols)
    apply_table_outline(ws)

    apply_copy_path_hyperlink_columns(
        ws,
        copy_cols,
        start_row=next_row,
        records=records,
        anchor_col_idx=category_col_idx,
    )
    apply_file_link_hyperlinks(ws, desired_keys, start_row=next_row)
    apply_hidden_tracking_url_columns(
        ws, records, start_row=next_row, vba_friendly=vba_friendly
    )
    apply_hidden_tracking_number_columns(
        ws, records, start_row=next_row, vba_friendly=vba_friendly
    )
    apply_open_tracking_list_column(
        ws, desired_keys, records, start_row=next_row, vba_friendly=vba_friendly
    )
    apply_open_tracking_numbers_web_column(
        ws, desired_keys, records, start_row=next_row, vba_friendly=vba_friendly,
    )
    apply_open_tracking_numbers_order_column(
        ws, desired_keys, records, start_row=next_row, vba_friendly=vba_friendly,
    )

    full_records = load_excel_records(PROJECT_ROOT, include_automation_hub=True, sync_pod_json=False)
    hub_record, full_data_records = _split_automation_hub_record(full_records)
    apply_gift_invoice_link_columns(ws, desired_keys, full_data_records, DATA_START_ROW, PROJECT_ROOT)
    apply_debug_gated_tracking_tool_columns(ws, desired_keys)
    if hub_record is not None:
        apply_special_row_styles(ws, start_row=ACTION_ROW, records=[hub_record], column_keys=desired_keys)
    apply_special_row_styles(ws, start_row=DATA_START_ROW, records=full_data_records, column_keys=desired_keys)

    center_invoice_and_shipping_headers(ws, desired_keys)

    ws.freeze_panes = FREEZE_PANES_CELL

    if Path(path).suffix.lower() == ".xlsm" and "Orders" in wb.sheetnames:
        macro_mod = _macro_template_module() if sys.platform == "win32" else None
        if macro_mod is not None:
            script_path = Path(__file__).resolve().parent / "copy_email_path_to_clipboard.py"
            ini_written = macro_mod.write_clipboard_launch_ini(
                CLIPBOARD_LAUNCH_INI_PATH,
                sys.executable,
                script_path,
                viewer_script=TRACKING_VIEWER_SCRIPT,
                giftcard_link_script=GIFT_INVOICE_LINK_SCRIPT,
                tracking_numbers_viewer_script=TRACKING_NUMBERS_VIEWER_SCRIPT,
                tracking_status_viewer_script=TRACKING_STATUS_VIEWER_SCRIPT,
                pod_workflow_script=POD_WORKFLOW_SCRIPT,
            )
            set_clipboard_ini_cell(wb, ini_written)

    wb.save(path)
    print(f"Appended {len(records)} row(s) to '{path}'.")


def reset_duplicate_flags(json_path: str):
    """Reset all duplicate_on_last_run flags to 0 in the JSON file."""
    with open(json_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    for record in records:
        record["duplicate_on_last_run"] = 0

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    print(f"Reset duplicate_on_last_run to 0 for {len(records)} record(s).")


def _remove_orders_template_for_rebuild() -> None:
    """Only remove the template when explicitly requested via env override."""
    if sys.platform != "win32":
        return
    force = (os.getenv("FORCE_REBUILD_ORDERS_TEMPLATE") or "").strip().lower()
    if force not in ("1", "true", "yes", "on"):
        return
    p = ORDERS_TEMPLATE_PATH
    try:
        if p.is_file():
            p.unlink()
            print(f"[createExcelDocument] Removed template for rebuild: {p}")
    except OSError as e:
        print(f"[createExcelDocument] Could not remove template ({e}); continuing.")


def main():
    _remove_orders_template_for_rebuild()

    records = load_excel_records(PROJECT_ROOT, include_automation_hub=True, sync_pod_json=True)
    _prefetch_17track_for_excel_build(records, quiet=False)

    wb, using_template = _build_populated_orders_workbook(records, verbose_clipboard_log=True)

    out_path = _resolve_excel_output_path(using_template)

    wb.save(out_path)
    print(f"Wrote '{out_path}' with {len(records)} row(s).")
    if not using_template:
        print(
            "Note: No macro template - output is .xlsx; Open File Location cells open the file. "
            "On Windows, install pywin32 + Excel so the program can auto-create "
            f"'{ORDERS_TEMPLATE_PATH}', or add that file manually (CLIPBOARD_SETUP.txt)."
        )

    record_excel_build_debug_mode()
    reset_duplicate_flags(JSON_PATH)


if __name__ == "__main__":
    print(f"\n{'='*60}")
    print(f"[createExcelDocument] Run started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")

    try:
        main()
        print("Excel creation finished successfully.")
    except Exception as e:
        print(f"\nERROR: {console_safe_text(e)}")
        sys.exit(1)
