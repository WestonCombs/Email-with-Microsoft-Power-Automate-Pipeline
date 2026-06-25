from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

_TMP = tempfile.TemporaryDirectory()
os.environ.setdefault("BASE_DIR", _TMP.name)

from openpyxl import Workbook  # noqa: E402

from createExcelDocument import createExcelDocument as excel_doc  # noqa: E402


class ExcelAccountingColumnTests(unittest.TestCase):
    @classmethod
    def tearDownClass(cls) -> None:
        _TMP.cleanup()

    def test_populate_orders_sheet_adds_accounting_column_and_dropdown(self) -> None:
        wb = Workbook()
        excel_doc.populate_orders_sheet(
            wb,
            [
                {
                    "email_category": "Invoice",
                    "order_number": "1234",
                    "purchase_datetime": "2026-05-22",
                    "company": "Walgreens",
                    "total_amount_paid": 10.0,
                    "subtotal_amount": 9.0,
                    "tax_paid": 1.0,
                    "excel_flagged": True,
                }
            ],
        )
        ws = wb["Orders"]
        keys, _labels = excel_doc._build_column_order()
        headers = [ws.cell(row=excel_doc.HEADER_ROW, column=i).value for i in range(1, len(keys) + 1)]
        self.assertEqual(headers[:3], ["Flagged", "Order Number", "Category"])
        self.assertNotIn("Shipping Status", headers)
        self.assertEqual(ws.freeze_panes, "D3")
        self.assertEqual(ws.auto_filter.ref, f"A2:{excel_doc.get_column_letter(ws.max_column)}{ws.max_row}")
        self.assertEqual(
            ws.cell(row=excel_doc.ACTION_ROW, column=keys.index("total_amount_paid") + 1).value,
            "Process Remaining PODs",
        )
        self.assertEqual(ws.cell(row=excel_doc.DATA_START_ROW, column=1).value, "True")
        accounting_col = keys.index("accounting") + 1
        subtotal_col = keys.index("subtotal_amount") + 1
        total_col = keys.index("total_amount_paid") + 1
        tax_col = keys.index("tax_paid") + 1
        gift_col = keys.index("gift_card_amount") + 1
        self.assertEqual(subtotal_col, total_col + 1)
        self.assertEqual(tax_col, subtotal_col + 1)
        self.assertEqual(gift_col, tax_col + 1)
        self.assertEqual(accounting_col, gift_col + 1)
        self.assertEqual(ws.cell(row=excel_doc.HEADER_ROW, column=subtotal_col).value, "Subtotal")
        self.assertEqual(ws.cell(row=excel_doc.DATA_START_ROW, column=subtotal_col).value, 9.0)
        self.assertEqual(ws.cell(row=excel_doc.HEADER_ROW, column=accounting_col).value, "Accounting")
        self.assertIsNone(ws.cell(row=excel_doc.DATA_START_ROW, column=accounting_col).value)
        self.assertEqual(
            ws.cell(row=excel_doc.HEADER_ROW, column=excel_doc.RECORD_ID_COL).value,
            "Record ID",
        )
        self.assertTrue(ws.column_dimensions[excel_doc.get_column_letter(excel_doc.RECORD_ID_COL)].hidden)
        self.assertTrue(str(ws.cell(row=excel_doc.DATA_START_ROW, column=excel_doc.RECORD_ID_COL).value).startswith("ord_"))
        validations = list(ws.data_validations.dataValidation)
        self.assertTrue(
            any(f"{excel_doc.get_column_letter(accounting_col)}3" in str(dv.sqref) for dv in validations)
        )
        self.assertTrue(any(dv.formula1 == '"Complete,Review"' for dv in validations))
        self.assertTrue(
            any("A3" in str(dv.sqref) and dv.formula1 == '"True,False"' for dv in validations)
        )
        category_col = keys.index("email_category") + 1
        self.assertTrue(
            any(
                f"{excel_doc.get_column_letter(category_col)}3" in str(dv.sqref)
                and dv.formula1 == '"Delivered,Invoice,Shipped,Gift Card,Unknown,POD"'
                for dv in validations
            )
        )

    def test_order_category_blocks_are_grouped_under_first_row(self) -> None:
        wb = Workbook()
        excel_doc.populate_orders_sheet(
            wb,
            [
                {"email_category": "Invoice", "order_number": "1234", "company": "A"},
                {"email_category": "Invoice", "order_number": "1234", "company": "A"},
                {"email_category": "Shipped", "order_number": "1234", "company": "A"},
                {"email_category": "Shipped", "order_number": "1234", "company": "A"},
                {"email_category": "Invoice", "order_number": "1234", "company": "A"},
                {"email_category": "Invoice", "order_number": "5678", "company": "B"},
            ],
        )
        ws = wb["Orders"]

        self.assertFalse(ws.sheet_properties.outlinePr.summaryBelow)
        self.assertEqual(ws.row_dimensions[excel_doc.DATA_START_ROW].outlineLevel, 0)
        self.assertEqual(ws.row_dimensions[excel_doc.DATA_START_ROW + 1].outlineLevel, 1)
        self.assertEqual(ws.row_dimensions[excel_doc.DATA_START_ROW + 2].outlineLevel, 0)
        self.assertEqual(ws.row_dimensions[excel_doc.DATA_START_ROW + 3].outlineLevel, 1)
        self.assertEqual(ws.row_dimensions[excel_doc.DATA_START_ROW + 4].outlineLevel, 0)
        self.assertEqual(ws.row_dimensions[excel_doc.DATA_START_ROW + 5].outlineLevel, 0)

    def test_unknown_purchase_date_cell_is_blank(self) -> None:
        wb = Workbook()
        excel_doc.populate_orders_sheet(
            wb,
            [
                {
                    "email_category": "Invoice",
                    "order_number": "1234",
                    "purchase_datetime": None,
                    "company": "Walgreens",
                }
            ],
        )
        ws = wb["Orders"]
        keys, _labels = excel_doc._build_column_order()
        purchase_col = keys.index("purchase_datetime") + 1
        self.assertIsNone(ws.cell(row=excel_doc.DATA_START_ROW, column=purchase_col).value)

    def test_migrate_accounting_column_preserves_shifted_cells(self) -> None:
        wb = Workbook()
        ws = wb.active
        _keys, labels = excel_doc._build_column_order()
        old_labels = [label for label in labels if label != "Accounting"]
        ws.append([""] * len(old_labels))
        ws.append(old_labels)
        row = [None] * len(old_labels)
        row[old_labels.index("Category")] = "Invoice"
        row[old_labels.index("Order Number")] = "1234"
        row[old_labels.index("Purchase Date")] = "2026-05-22"
        row[old_labels.index("Company")] = "Walgreens"
        row[old_labels.index("Total Paid")] = 10
        row[old_labels.index("Subtotal")] = 9
        row[old_labels.index("Tax Paid")] = 1
        row[old_labels.index("Invoice Link")] = "Invoice Link"
        ws.append(row)
        ws.cell(row=excel_doc.DATA_START_ROW, column=excel_doc.COPY_PATH_URI_COL, value="file:///C:/tmp/invoice.pdf")

        excel_doc._migrate_accounting_column_if_missing(ws, _keys)

        accounting_col = _keys.index("accounting") + 1
        self.assertEqual(ws.cell(row=excel_doc.HEADER_ROW, column=accounting_col).value, "Accounting")
        self.assertIsNone(ws.cell(row=excel_doc.DATA_START_ROW, column=accounting_col).value)
        self.assertEqual(ws.cell(row=excel_doc.DATA_START_ROW, column=accounting_col + 1).value, "Invoice Link")
        self.assertEqual(
            ws.cell(row=excel_doc.DATA_START_ROW, column=excel_doc.COPY_PATH_URI_COL).value,
            "file:///C:/tmp/invoice.pdf",
        )


if __name__ == "__main__":
    unittest.main()
